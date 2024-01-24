import openpyxl

workbook = openpyxl.load_workbook("E:\Case01.xlsx")  # 打开Excel文件，仅支持xlsx格式
sheet = workbook['登录']  # 打开要要操作的sheet表单
print(sheet.cell(2, 3).value)  # 读取当前sheet页，单元格第2行、第3列的值
res = sheet.cell(2, 3).value = "登录成功"  # 修改当前sheet页，单元格第2行、第3列的值，将值改为 登录成功？
workbook.save("E:\Case01.xlsx")  # 保存操作
print(res)
